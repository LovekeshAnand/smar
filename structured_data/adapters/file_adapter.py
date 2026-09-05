"""
structured_data/adapters/file_adapter.py
========================================
High-performance Tabular File Storage Adapter for SMAR v2.
Supports CSV, TSV, and Excel (.xlsx, .xls) files.
Dynamically introspects tabular headers, creates an in-memory or on-disk
indexed SQLite backing database with FTS5 search, and provides unified sub-millisecond querying.
"""

import os
import csv
import sqlite3
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path

from .base import BaseStorageAdapter

logger = logging.getLogger("smar.structured_data.adapters.file")

VOLATILE_FIELDS = {"quantity", "stock", "qty", "unit_price", "price", "mrp", "cost", "cost_price", "rate", "bhav", "is_active", "updated_at"}


class FileStorageAdapter(BaseStorageAdapter):
    """
    Adapter for CSV, TSV, and Excel spreadsheet files.
    Ingests and indexes file data on-the-fly into an optimized SQLite store.
    """

    def __init__(self, file_path: str, primary_key: Optional[str] = None):
        self.file_path = os.path.abspath(file_path)
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File not found: {self.file_path}")

        self.ext = Path(self.file_path).suffix.lower()
        self.table_name = "file_records"
        self.primary_key = primary_key

        # In-memory SQLite connection for rapid indexing
        self.conn = sqlite3.connect(":memory:", check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA synchronous = OFF;")
        self.conn.execute("PRAGMA journal_mode = MEMORY;")

        self.columns: List[str] = []
        self._load_and_index_file()

    def get_source_name(self) -> str:
        return f"File ({os.path.basename(self.file_path)})"

    def get_source_type(self) -> str:
        if self.ext in [".xlsx", ".xls"]:
            return "excel"
        return "csv"

    def _read_rows(self) -> List[Dict[str, Any]]:
        """Reads rows from CSV or Excel file."""
        rows: List[Dict[str, Any]] = []

        if self.ext in [".xlsx", ".xls"]:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(self.file_path, data_only=True)
                sheet = wb.active
                iter_rows = sheet.iter_rows(values_only=True)
                headers = [str(h).strip() for h in next(iter_rows) if h is not None]
                for r in iter_rows:
                    if not any(r):
                        continue
                    row_dict = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        row_dict[h] = val
                    rows.append(row_dict)
                wb.close()
            except ImportError:
                import pandas as pd
                df = pd.read_excel(self.file_path)
                rows = df.to_dict(orient="records")
        else:
            # CSV / TSV
            delimiter = "\t" if self.ext == ".tsv" else ","
            with open(self.file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                for r in reader:
                    rows.append(dict(r))

        return rows

    def _load_and_index_file(self) -> None:
        """Loads file rows into indexed in-memory SQLite store."""
        raw_rows = self._read_rows()
        if not raw_rows:
            logger.warning(f"File '{self.file_path}' has 0 rows.")
            return

        # Determine column names (clean whitespace and lowercased identifier)
        sample_row = raw_rows[0]
        self.columns = [k.strip() for k in sample_row.keys() if k and k.strip()]

        # Identify Primary Key candidate if not explicitly given
        if not self.primary_key:
            pk_candidates = ["item_id", "id", "sku", "code", "barcode", "product_id"]
            for cand in pk_candidates:
                if cand in self.columns:
                    self.primary_key = cand
                    break
            if not self.primary_key:
                self.primary_key = self.columns[0]

        # Infer basic data types from sample values
        col_definitions = []
        for col in self.columns:
            sample_val = sample_row.get(col)
            col_type = "TEXT"
            if isinstance(sample_val, (int, float)):
                col_type = "REAL"
            col_definitions.append(f'"{col}" {col_type}')

        # Create Table
        ddl = f"CREATE TABLE {self.table_name} (\n  " + ",\n  ".join(col_definitions) + "\n);"
        self.conn.execute(ddl)

        # Batch Insert
        placeholders = ", ".join(["?"] * len(self.columns))
        col_names = ", ".join([f'"{c}"' for c in self.columns])
        insert_sql = f"INSERT INTO {self.table_name} ({col_names}) VALUES ({placeholders})"

        data_tuples = []
        for r in raw_rows:
            t = tuple(r.get(c) for c in self.columns)
            data_tuples.append(t)

        self.conn.executemany(insert_sql, data_tuples)
        self.conn.commit()

        # Build Indexes
        # 1. Primary Key Index
        if self.primary_key:
            self.conn.execute(f'CREATE INDEX idx_file_pk ON {self.table_name} ("{self.primary_key}");')

        # 2. Text / Name Index
        for cand_name in ["canonical_name", "name", "item_name", "product_name", "title", "normalized_name"]:
            if cand_name in self.columns:
                self.conn.execute(f'CREATE INDEX idx_file_{cand_name} ON {self.table_name} ("{cand_name}");')

        # 3. FTS5 Virtual Table over text columns
        text_cols = [c for c in self.columns if c.lower() not in VOLATILE_FIELDS and not c.lower().endswith(('_id', '_dt', 'date'))][:4]
        if text_cols:
            fts_cols_str = ", ".join([f'"{c}"' for c in text_cols])
            self.conn.execute(f"CREATE VIRTUAL TABLE file_fts USING fts5({fts_cols_str}, content='{self.table_name}');")
            self.conn.execute(f"INSERT INTO file_fts(rowid, {fts_cols_str}) SELECT rowid, {fts_cols_str} FROM {self.table_name};")
            self.conn.commit()

        logger.info(f"Loaded {len(raw_rows):,} rows from '{self.file_path}' into indexed in-memory store.")

    def introspect_schema(self) -> Dict[str, Any]:
        cursor = self.conn.cursor()
        cols_info = cursor.execute(f"PRAGMA table_info({self.table_name});").fetchall()
        row_cnt = cursor.execute(f"SELECT COUNT(*) FROM {self.table_name};").fetchone()[0]

        columns = []
        for col in cols_info:
            col_name = col[1]
            col_type = col[2] or "TEXT"
            is_pk = (col_name == self.primary_key)
            is_volatile = col_name.lower() in VOLATILE_FIELDS

            sample_vals = []
            try:
                sample_vals = [r[0] for r in cursor.execute(f'SELECT DISTINCT "{col_name}" FROM {self.table_name} WHERE "{col_name}" IS NOT NULL LIMIT 3;').fetchall()]
            except Exception:
                pass

            columns.append({
                "name": col_name,
                "type": col_type,
                "is_primary_key": is_pk,
                "is_volatile": is_volatile,
                "sample_values": sample_vals
            })

        return {
            "source_name": self.get_source_name(),
            "source_type": self.get_source_type(),
            "tables": [
                {
                    "table_name": self.table_name,
                    "columns": columns,
                    "row_count": row_cnt,
                    "indexes": [f"idx_file_{self.primary_key}"]
                }
            ]
        }

    def get_item_by_id(self, item_id: str, table_name: Optional[str] = None) -> Optional[Dict[str, Any]]:
        if not self.primary_key:
            return None
        cursor = self.conn.cursor()
        row = cursor.execute(f'SELECT * FROM {self.table_name} WHERE "{self.primary_key}" = ?', (item_id,)).fetchone()
        return dict(row) if row else None

    def search_by_text(self, query: str, limit: int = 20, table_name: Optional[str] = None) -> List[Dict[str, Any]]:
        cursor = self.conn.cursor()
        # Try FTS5 first
        try:
            tokens = [f"{t}*" for t in query.strip().split() if t]
            if tokens:
                match_str = " ".join(tokens)
                sql = f"""
                SELECT r.* 
                FROM file_fts f 
                JOIN {self.table_name} r ON f.rowid = r.rowid 
                WHERE file_fts MATCH ? 
                LIMIT ?;
                """
                rows = cursor.execute(sql, (match_str, limit)).fetchall()
                if not rows and len(tokens) > 1:
                    match_str_or = " OR ".join(tokens)
                    rows = cursor.execute(sql, (match_str_or, limit)).fetchall()
                if rows:
                    return [dict(r) for r in rows]
        except Exception:
            pass

        # Fallback to column LIKE
        name_cols = [c for c in self.columns if "name" in c.lower() or "item" in c.lower() or "desc" in c.lower()]
        target_col = name_cols[0] if name_cols else self.columns[0]
        pattern = f"%{query.strip()}%"
        rows = cursor.execute(f'SELECT * FROM {self.table_name} WHERE "{target_col}" LIKE ? LIMIT ?', (pattern, limit)).fetchall()
        return [dict(r) for r in rows]

    def filter_items(
        self,
        filters: Dict[str, Any],
        limit: int = 50,
        table_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        cursor = self.conn.cursor()
        where_clauses = []
        params = []
        for col, val in filters.items():
            if col in self.columns:
                where_clauses.append(f'"{col}" = ?')
                params.append(val)

        where_str = " AND ".join(where_clauses) if where_clauses else "1=1"
        sql = f"SELECT * FROM {self.table_name} WHERE {where_str} LIMIT ?"
        params.append(limit)

        rows = cursor.execute(sql, tuple(params)).fetchall()
        return [dict(r) for r in rows]

    def get_aggregations(
        self,
        group_by: Optional[str] = None,
        table_name: Optional[str] = None
    ) -> Dict[str, Any]:
        cursor = self.conn.cursor()
        total_cnt = cursor.execute(f"SELECT COUNT(*) FROM {self.table_name};").fetchone()[0]
        res = {"total_records": total_cnt}

        if group_by and group_by in self.columns:
            rows = cursor.execute(f'SELECT "{group_by}", COUNT(*) as cnt FROM {self.table_name} GROUP BY "{group_by}" ORDER BY cnt DESC LIMIT 20;').fetchall()
            res[f"by_{group_by}"] = {r[0]: r[1] for r in rows}

        return res

    def get_total_count(self, table_name: Optional[str] = None) -> int:
        cursor = self.conn.cursor()
        row = cursor.execute(f"SELECT COUNT(*) FROM {self.table_name};").fetchone()
        return row[0] if row else 0

    def close(self) -> None:
        try:
            self.conn.close()
        except Exception:
            pass
